import os
import sys
import numpy as np
import pandas as pd
import requests
import html  # To handle HTML entities in abstracts
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError
import re
import time
import xml.etree.ElementTree as ET # Import ElementTree
from openpyxl import load_workbook

# Initialize the Nominatim geocoder
geolocator = Nominatim(user_agent="affiliation_splitter")

def geocode_address(address):
    """
    Tries to geocode the address using Nominatim API.
    Returns True if the address is valid, otherwise False.
    """
    time.sleep(1)  # Add a delay of 1 second between requests
    try:
        location = geolocator.geocode(address)
        return location is not None
    except (GeocoderTimedOut, GeocoderServiceError):
        return False

def separate_affiliation(affiliation):
    """
    Separates the affiliation into two parts:
    1. Affiliation Name: The name of the research team and university.
    2. Affiliation Address: The address part, validated via geocoding.
    """
    print(affiliation)
    affiliation = affiliation.replace(";", " ")
    parts = affiliation.split(',')

    # Regular expression pattern for common terms in affiliation names
    name_keywords_pattern = re.compile(r"(universit|institut|college|department|facult|school|academy|center|centre|"
                                      r"lab|laboratory|division|research|unit|corporation)", re.IGNORECASE)

    # Try to separate based on geocoding the latter part of the affiliation
    for i in range(1, len(parts)):
        potential_address = ','.join(parts[i:]).strip()
        potential_name = ','.join(parts[:i]).strip()
        # Check if the potential address contains any keywords typically found in a name
        if not name_keywords_pattern.search(potential_address):
            if geocode_address(potential_address):
                return potential_name, potential_address

    # If geocoding fails, return the full affiliation as name and empty address
    return affiliation, ""

def get_orcid_data(orcid_id):
    url = f"https://pub.orcid.org/v3.0/{orcid_id}/person"
    headers = {'Accept': 'application/json'}

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()

        name = data.get('name', {})
        given_names = name.get('given-names', {}).get('value', '')

        aff = data.get('employments', {}).get('affiliation-group', [{}])[0]
        aff_summary = aff.get('summaries', [{}])[0].get('employment-summary', {})
        institution = aff_summary.get('organization', {}).get('name', '')

        return given_names, institution
    except requests.exceptions.RequestException:
        return None, None

def map_publication_type(crossref_type):
    """Maps CrossRef type of publication to the Excel format."""
    type_mapping = {
        "journal-article": "journal",
        "journal": "journal",
        "journal-volume": "journal",
        "journal-issue": "journal",
        "book": "book",
        "edited-book": "book",
        "book-series": "book",
        "book-set": "book",
        "book-chapter": "chapter",
        "book-section": "chapter",
        "book-part": "chapter",
        "report": "report",
        "report-component": "report",
        "report-series": "report",
        "proceedings": "conference",
        "proceedings-article": "conference",
        "proceedings-series": "conference"
    }
    return type_mapping.get(crossref_type, "other")

def fetch_crossref_metadata(doi, excel_ontology_file, index):
    url = f"https://api.crossref.org/works/{doi}"
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise exception for HTTP errors
        data = response.json().get('message', {})

        # Extracting metadata
        if not excel_ontology_file.loc[index, ("Paper metadata", "Title")] or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Title")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Title")] = data.get('title', [''])[0]
        if not excel_ontology_file.loc[index, ("Paper metadata", "type of publication")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "type of publication")]):
            excel_ontology_file.loc[index, ("Paper metadata", "type of publication")] = map_publication_type( data.get('type', '') )
        if not excel_ontology_file.loc[index, ("Paper metadata", "journal")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "journal")]):
            excel_ontology_file.loc[index, ("Paper metadata", "journal")] = data.get('container-title', [''])[0]
        if not excel_ontology_file.loc[index, ("Paper metadata", "Year")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Year")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Year")] = data.get('created', {}).get('date-parts', [['']])[0][0]
        if not excel_ontology_file.loc[index, ("Paper metadata", "Keywords")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Keywords")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Keywords")] =  " ; ".join(data.get('subject', []))
        if not excel_ontology_file.loc[index, ("Paper metadata", "Abstract")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Abstract")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Abstract")] = html.unescape(data.get('abstract', '')) if data.get('abstract') else ''
        if not excel_ontology_file.loc[index, ("Paper metadata", "number of citations")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "number of citations")]):
            excel_ontology_file.loc[index, ("Paper metadata", "number of citations")] = data.get('is-referenced-by-count', '')

        # Extracting authors
        authors_data = data.get('author', [])
        authors_names = []
        authors_affiliation_names = []
        authors_affiliation_addresses = []
        for author in authors_data:
            first_name = author.get('given', '')
            last_name = author.get('family', '')
            ORCID = author.get('ORCID', '')
            affiliation = author.get('affiliation', [{'name': ''}])
            if not affiliation or len(affiliation)==0:
                affiliation_name = ""
                affiliation_address = ""
            else:
                affiliation_name, affiliation_address = separate_affiliation(affiliation[0].get('name', ''))

            #if ORCID provided and if it lacks one part of the affiliation or if the first name is abbreviated, use orcid API to complete the information
            if ORCID and (not first_name or '.' in first_name or not affiliation_address):
                print(ORCID)
                orcid_first_name, orcid_affiliation = get_orcid_data(ORCID.split('/')[-1])
                first_name = orcid_first_name if (orcid_first_name and (not first_name or '.' in first_name)) else first_name
                if orcid_affiliation and not affiliation_address:
                    affiliation_name, affiliation_address = separate_affiliation(orcid_affiliation)

            authors_affiliation_names.append(affiliation_name)
            authors_affiliation_addresses.append(affiliation_address)
            print(first_name, last_name)
            if first_name:
                authors_names.append(last_name+", "+first_name)
            else:
                authors_names.append(last_name)
        print(authors_names)
        print(" and ".join(authors_names))
        if not excel_ontology_file.loc[index, ("Paper metadata", "Authors")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Authors")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Authors")] = " and ".join(authors_names)
            print("crossref1", excel_ontology_file.loc[index, ("Paper metadata", "Authors")])
        if not excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Name")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Name")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Name")] = " ; ".join(authors_affiliation_names)
        if not excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Address")]  or pd.isna(excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Address")]):
            excel_ontology_file.loc[index, ("Paper metadata", "Affiliation Address")] = " ; ".join(authors_affiliation_addresses)
        return excel_ontology_file

    except requests.exceptions.RequestException as e:
        print(f"Error fetching DOI data: {e}")
        return excel_ontology_file

def fetch_doaj_metadata(doi, excel_ontology_file, index):
    """
    Fetch metadata from the DOAJ API and update the Excel ontology file.
    Only updates fields that are currently empty.
    """
    doaj_url = f"https://doaj.org/api/v2/search/articles/doi:{doi}"

    try:
        response = requests.get(doaj_url)
        response.raise_for_status()
        data = response.json()

        if data.get('total', 0) > 0:
            article = data['results'][0]
            title = article['bibjson'].get('title', '')
            abstract = article['bibjson'].get('abstract', '')
            keywords = article['bibjson'].get('keywords', [])

            # Update the Excel file only if fields are empty
            if not excel_ontology_file.loc[index, ("Paper metadata", "Title")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Title")] = title
            if not excel_ontology_file.loc[index, ("Paper metadata", "Abstract")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Abstract")] = abstract
            if not excel_ontology_file.loc[index, ("Paper metadata", "Keywords")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Keywords")] = " ; ".join(keywords)

        return excel_ontology_file

    except requests.exceptions.RequestException as e:
        print(f"Error fetching from DOAJ: {e}")
        return excel_ontology_file


def fetch_arxiv_metadata(doi, excel_ontology_file, index):
    """
    Fetch metadata from the arXiv API and update the Excel ontology file.
    Only updates fields that are currently empty.
    """
    arxiv_url = f"http://export.arxiv.org/api/query?search_query=doi:{doi}&max_results=1"

    try:
        response = requests.get(arxiv_url)
        response.raise_for_status()

        # Parse the XML response
        root = ET.fromstring(response.content)

        # Namespace for arXiv metadata
        ns = {'arxiv': 'http://www.w3.org/2005/Atom'}

        # Extract the entry
        entry = root.find('arxiv:entry', ns)

        if entry is not None:
            title = entry.find('arxiv:title', ns).text.strip() if entry.find('arxiv:title', ns) is not None else ''
            abstract = entry.find('arxiv:summary', ns).text.strip() if entry.find('arxiv:summary', ns) is not None else ''
            authors = [author.text.strip() for author in entry.findall('arxiv:author/arxiv:name', ns)]

            # Update the Excel file only if fields are empty
            if not excel_ontology_file.loc[index, ("Paper metadata", "Title")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Title")] = title
            if not excel_ontology_file.loc[index, ("Paper metadata", "Abstract")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Abstract")] = abstract
            if not excel_ontology_file.loc[index, ("Paper metadata", "Authors")]:
                excel_ontology_file.loc[index, ("Paper metadata", "Authors")] = " and ".join(authors)

        return excel_ontology_file

    except requests.exceptions.RequestException as e:
        print(f"Error fetching from arXiv: {e}")
        return excel_ontology_file

def enrich_metadata(excel_ontology_file):
    for i in range(1, len(excel_ontology_file)):
        doi = excel_ontology_file.loc[i, ("Paper metadata", "doi")]

        if (doi=="") or pd.isna(doi):
            #It is either the same paper as the row above, or a paper without doi
            #If there is a title which is not the same as above, or that it is the first row, we can suppose it is a paper without doi
            title_i = excel_ontology_file.loc[i, ("Paper metadata", "Title")]

            if (i==1) or ( not ((title_i=="") or pd.isna(title_i)) and (excel_ontology_file.loc[i-1, ("Paper metadata", "Title")]!=title_i)):
                #It is an article without doi. Thus, we don't try to get the metatada
                excel_ontology_file.loc[i, ("Paper metadata", "doi")] = "NoDoi_"+str(time.time())
            else:
                #It is likely to be the same article as above
                #We may copy all the previous row paper metadata
                excel_ontology_file.loc[i]["Paper metadata"] = excel_ontology_file.loc[i-1, "Paper metadata"]
            continue
        else:
            doi = doi.replace("https://doi.org/", "")
            excel_ontology_file.loc[i, ("Paper metadata", "doi")] = doi
        print("already existing names", excel_ontology_file.loc[i, ("Paper metadata", "Authors")])
        excel_ontology_file = fetch_crossref_metadata(doi, excel_ontology_file, i)
        print("crossref2", excel_ontology_file.loc[i, ("Paper metadata", "Authors")])
        excel_ontology_file = fetch_doaj_metadata(doi, excel_ontology_file, i)
        print("doaj", excel_ontology_file.loc[i, ("Paper metadata", "Authors")])
        excel_ontology_file = fetch_arxiv_metadata(doi, excel_ontology_file, i)
        print("arxiv", excel_ontology_file.loc[i, ("Paper metadata", "Authors")])
    return excel_ontology_file

#%%
if __name__ == "__main__":

    #input path
    if len(sys.argv)>1:
        excel_ontology_file_path = os.path.join(sys.argv[1])
    else:
        excel_ontology_file_path = os.path.join(
                "LULC_Ontology_example.xlsm"
                )

    #output path
    if len(sys.argv)>2:
        output_path = os.path.join(sys.argv[1])
    else:
        output_path = excel_ontology_file_path

    excel_ontology_file = pd.read_excel(excel_ontology_file_path, dtype=str, sheet_name="ontology_instanciation", header=[0, 1])
    excel_ontology_file = enrich_metadata(excel_ontology_file)

    #saving the excel file with the same display
    wb = load_workbook(file_path, keep_vba=True)
    ws = wb["ontology_instanciation"]

    # write values manually (preserves layout)
    for i, row in enumerate(df.values, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    wb.save(output_path)
